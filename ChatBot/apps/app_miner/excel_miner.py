import re
import openpyxl
import datetime
import traceback
import logging
import json
import time
from openpyxl.cell.cell import MergedCell
from django.db import transaction

# Import các Model đã cập nhật
from .models import ExcelSheet, DataField, ExcelTableRegion
from apps.app_knowledge.models import KnowledgeDraft
from apps.app_coach.models import DataType

logger = logging.getLogger(__name__)

class ExcelMinerService:
    def __init__(self):
        # Danh sách các từ khóa để nhận diện UI (Nút bấm, rác)
        self.ui_keywords = [
            'lưu', 'tìm kiếm', 'xóa', 'thêm', 'in phiếu', 'thoát', 'đóng', 
            'update', 'filter', 'search', 'edit', 'save', 'print'
        ]
        # Màu sắc quy ước cho các trường bắt buộc (Tiệm vàng thường dùng)
        self.required_colors = ["FFFF0000", "FFFFC000"] # Đỏ, Cam

    @classmethod
    def run_workflow(cls, project):
        miner = cls()
        project.status = 'PROCESSING'
        project.save()
        
        success, message = miner.process_project(project)
        
        project.status = 'COMPLETED' if success else 'FAILED'
        project.save()
        return success, message

    def process_project(self, project):
        try:
            # Load workbook (data_only=False để lấy công thức)
            wb = openpyxl.load_workbook(project.file_path.path, data_only=False)
            total_sheets = len(wb.sheetnames)

            for index, sheet_name in enumerate(wb.sheetnames):
                ws = wb[sheet_name]
                print(f"--- [Miner] Đang bóc tách Sheet: {sheet_name} ({index+1}/{total_sheets}) ---")
                
                # 1. Khởi tạo/Cập nhật ExcelSheet
                sheet_obj, _ = ExcelSheet.objects.update_or_create(
                    project=project, 
                    name=sheet_name,
                    defaults={'refine_status': 'PENDING'} 
                )

                # Dọn dẹp dữ liệu cũ để tránh trùng lặp
                DataField.objects.filter(sheet=sheet_obj).delete()
                
                # 2. Container để gom Metadata đã phân loại
                sheet_intel = {
                    "logic_blocks": [],   # Công thức tính toán
                    "ui_elements": [],    # Nút bấm, popup, nhãn giao diện
                    "business_data": [],  # Các trường dữ liệu nghiệp vụ chính
                    "raw_structure": []   # Cấu trúc thô để AI tham khảo vị trí
                }

                fields_to_create = []
                merged_ranges = ws.merged_cells.ranges

                # 3. Duyệt Sheet và Phân loại
                for row in ws.iter_rows(max_row=ws.max_row, max_col=ws.max_column):
                    if not any(cell.value is not None for cell in row):
                        continue 

                    for cell in row:
                        # Bỏ qua các ô phụ trong vùng Merged
                        if any(cell.coordinate in r and cell.coordinate != r.start_cell.coordinate for r in merged_ranges):
                            continue

                        if cell.value is not None or (cell.comment):
                            # PHÂN LOẠI THÔ (Tư duy của anh Vũ)
                            field_type, label = self._classify_cell(ws, cell)
                            
                            # Chuẩn bị field object
                            field_obj = self._build_field_obj(sheet_obj, cell, field_type, label)
                            fields_to_create.append(field_obj)

                            # GOM VÀO METADATA TỔNG HỢP
                            item_info = {
                                "address": cell.coordinate,
                                "label": label,
                                "value": str(cell.value) if not field_obj.formula else "FORMULA",
                                "formula": field_obj.formula,
                                "comment": cell.comment.text if cell.comment else None
                            }

                            if field_type == 'LOGIC':
                                sheet_intel["logic_blocks"].append(item_info)
                            elif field_type == 'UI':
                                sheet_intel["ui_elements"].append(item_info)
                            elif field_type == 'DATA':
                                sheet_intel["business_data"].append(item_info)

                # 4. Lưu Batch DataFields vào SQL
                if fields_to_create:
                    DataField.objects.bulk_create(fields_to_create)

                # 5. Đóng gói "Cục quặng" vào Sheet Metadata
                sheet_obj.metadata = sheet_intel
                sheet_obj.refine_status = 'EXTRACTED'
                
                # Tạo tóm tắt sơ bộ để Admin nhìn thấy ngay
                sheet_obj.description = (
                    f"Sheet có {len(sheet_intel['logic_blocks'])} logic tính toán, "
                    f"{len(sheet_intel['business_data'])} trường nghiệp vụ. "
                    f"Đã nhận diện {len(sheet_intel['ui_elements'])} thành phần UI."
                )
                sheet_obj.save()

                # 6. Tạo một KnowledgeDraft DUY NHẤT cho cả Sheet (Thay vì 10k cái lẻ)
                self._create_unified_draft(project, sheet_obj)

            return True, f"Thành công! Đã xử lý {total_sheets} sheets."

        except Exception as e:
            logger.error(f"Lỗi Miner: {str(e)}")
            return False, traceback.format_exc()

    def _classify_cell(self, ws, cell):
        """Logic phân loại thông minh của anh Vũ"""
        val_str = str(cell.value).strip().lower() if cell.value is not None else ""
        
        # 1. Ưu tiên LOGIC (Nếu có dấu =)
        if val_str.startswith('='):
            return 'LOGIC', self._get_smart_label(ws, cell)

        # 2. Kiểm tra UI (Dựa trên keywords)
        if any(kw in val_str for kw in self.ui_keywords):
            return 'UI', cell.value

        # 3. Kiểm tra TRASH (Dữ liệu quá ngắn hoặc rác hệ thống)
        if len(val_str) > 100 and not cell.value: # Ví dụ
            return 'TRASH', "Long_Text_Garbage"

        # 4. Còn lại là DATA
        return 'DATA', self._get_smart_label(ws, cell)

    def _get_smart_label(self, ws, cell):
        """Tìm nhãn nghiệp vụ (Trái hoặc Trên)"""
        try:
            # Thử ô bên trái
            if cell.column > 1:
                l_val = ws.cell(row=cell.row, column=cell.column - 1).value
                if l_val and isinstance(l_val, str): return l_val.strip()
            # Thử ô bên trên
            if cell.row > 1:
                t_val = ws.cell(row=cell.row - 1, column=cell.column).value
                if t_val and isinstance(t_val, str): return t_val.strip()
        except: pass
        return f"Cell_{cell.coordinate}"

    def _build_field_obj(self, sheet_obj, cell, field_type, label):
        val_str = str(cell.value) if cell.value is not None else ""
        bg_color = "FFFFFF"
        if cell.fill and hasattr(cell.fill, 'start_color'):
            bg_color = str(cell.fill.start_color.rgb)

        return DataField(
            sheet=sheet_obj,
            cell_address=cell.coordinate,
            field_type=field_type,
            label=label[:255] if label else None,
            value=val_str if not val_str.startswith('=') else None,
            formula=val_str if val_str.startswith('=') else None,
            # color_code=bg_color
        )

    def _create_unified_draft(self, project, sheet_obj):
        """Tạo 1 bản thảo gom cho cả sheet thay vì lẻ tẻ"""
        KnowledgeDraft.objects.update_or_create(
            project=project,
            sheet=sheet_obj, # Liên kết trực tiếp với Sheet
            defaults={
                'term': f"Nghiệp vụ Sheet {sheet_obj.name}",
                'content': f"Bản thảo chờ AI phân tích nghiệp vụ cho sheet {sheet_obj.name}.",
                'category': 'SHEET_LOGIC',
                'status': 'PENDING',
                'origin_metadata': sheet_obj.metadata # Đẩy cục metadata đã phân loại qua đây
            }
        )