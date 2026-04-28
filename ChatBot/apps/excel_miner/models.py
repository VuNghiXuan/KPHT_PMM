import openpyxl
from django.db import models, transaction
import datetime 
from apps.system_monitor.models import DataType
from openpyxl.cell.cell import MergedCell
# Import DataType từ app system_monitor để dùng chung "từ điển"

class ExcelProject(models.Model):
    name = models.CharField(max_length=255, verbose_name="Tên dự án")
    file_path = models.FileField(upload_to='excels/', verbose_name="File gốc")
    uploaded_at = models.DateTimeField(auto_now_add=True)

    def save(self, *args, **kwargs):
        is_new = self.pk is None
        super().save(*args, **kwargs)
        
        # Chỉ chạy máy nghiền sau khi dự án đã thực sự nằm trong DB
        if is_new or self.file_path:
            # transaction.on_commit sẽ chờ DB lưu xong hoàn toàn mới chạy hàm bên trong
            transaction.on_commit(lambda: self.process_excel_file())

    def process_excel_file(self):
        """Hàm 'máy nghiền' tri thức: Phiên bản chịu tải cao cho 200+ sheets"""
        # Đảm bảo dữ liệu project đã được commit vào DB trước khi query ngược lại
        self.refresh_from_db()
        
        try:
            # Sử dụng read_only=False để đọc được cả Comments, nhưng tối ưu memory bằng cách load workbook một lần
            wb = openpyxl.load_workbook(self.file_path.path, data_only=False)
            
            # 1. Lấy từ điển nghiệp vụ từ DB
            from apps.system_monitor.models import DataType 
            data_types = DataType.objects.all()
            keywords_dict = {dt.name.lower(): dt.code for dt in data_types}
            
            # 2. Chuẩn bị các DataType mặc định
            default_dtype, _ = DataType.objects.get_or_create(code='MOCK', defaults={'name': 'Dữ liệu thô'})
            formula_dtype, _ = DataType.objects.get_or_create(code='FORMULA', defaults={'name': 'Công thức'})

            total_sheets = len(wb.sheetnames)
            print(f"🚀 Bắt đầu vét {total_sheets} sheets từ file: {self.name}")

            for index, sheet_name in enumerate(wb.sheetnames):
                ws = wb[sheet_name]
                
                # Print tiến độ để anh soi ở Terminal (biết ngay nếu dừng ở 12)
                print(f"--- [{index + 1}/{total_sheets}] Đang xử lý Sheet: {sheet_name}")

                # update_or_create để tránh lỗi Unique nếu anh chạy lại lần 2
                sheet_obj, _ = ExcelSheet.objects.update_or_create(
                    project=self, 
                    name=sheet_name,
                    defaults={'sheet_index': index, 'category': 'TAB_UI'}
                )

                # Xóa dữ liệu cũ của sheet này để nạp mới hoàn toàn
                DataField.objects.filter(sheet=sheet_obj).delete()
                
                fields_to_create = []
                
                # GIỚI HẠN VÙNG QUÉT: Chỉ quét tối đa 1000 dòng, 50 cột để tránh treo máy
                # File tiệm vàng thường không vượt quá ngưỡng này cho một biểu mẫu
                for row in ws.iter_rows(max_row=1000, max_col=50):
                    for cell in row:
                        # Kiểm tra xem ô có dữ liệu, công thức hoặc ghi chú không
                        if cell.value is not None or (hasattr(cell, 'comment') and cell.comment):
                            field_obj = self._prepare_cell_logic(
                                sheet_obj, cell, keywords_dict, default_dtype, formula_dtype
                            )
                            fields_to_create.append(field_obj)
                
                # 3. Lưu vào Database theo Batch (Mỗi lần 500 ô) để SQLite không bị nghẽn (Locked)
                if fields_to_create:
                    DataField.objects.bulk_create(fields_to_create, batch_size=500)
                    print(f"   ✅ Đã lưu {len(fields_to_create)} ô dữ liệu.")

            print(f"🔥 HOÀN THÀNH: Đã vét sạch {total_sheets} sheets cho dự án {self.name}.")

        except Exception as e:
            # In lỗi chi tiết để anh biết chính xác sheet nào gây lỗi
            import traceback
            print(f"❌ LỖI HỆ THỐNG tại sheet '{sheet_name if 'sheet_name' in locals() else 'N/A'}':")
            print(traceback.format_exc())

    def _prepare_cell_logic(self, sheet_obj, cell, keywords_dict, default_dtype, formula_dtype):
        """
        PHIÊN BẢN HOÀN CHỈNH: Xử lý triệt để MergedCell, Datetime, Date, Time và Metadata.
        """
        # 0. Xác định giá trị và xử lý ô gộp (Merged Cell)
        is_merged = isinstance(cell, MergedCell)
        raw_val = cell.value if not is_merged else None 
        
        # --- XỬ LÝ DATE/TIME CHO JSON SERIALIZATION ---
        # Ép kiểu tất cả các định dạng thời gian về chuỗi ISO để JSONField không bị crash
        if isinstance(raw_val, (datetime.datetime, datetime.date, datetime.time)):
            raw_val = raw_val.isoformat()
        # ----------------------------------------------

        val_str = str(raw_val).strip() if raw_val is not None else ""
        
        # 1. Chuẩn hóa chuỗi để so khớp từ điển (Xóa dấu cách thừa, viết thường)
        clean_val = " ".join(val_str.lower().split())

        # 2. Xác định UI Type & Định dạng màu sắc
        bg_color = "N/A"
        if hasattr(cell.fill, 'start_color'):
            # Lấy mã ARGB hoặc Index màu tùy theo loại màu trong Excel
            if cell.fill.start_color.type == 'indexed':
                bg_color = str(cell.fill.start_color.index)
            else:
                bg_color = str(cell.fill.start_color.rgb)

        ui_type = "DATA_CELL"
        is_bold = cell.font.bold if hasattr(cell.font, 'bold') else False
        
        # Logic nhận diện nhanh giao diện (Header thường in đậm và có màu)
        if is_bold and bg_color != "N/A" and bg_color != "00000000":
            ui_type = "GRID_HEADER"
        elif bg_color != "00000000" and bg_color != "N/A" and len(val_str) < 25 and not val_str.startswith('='):
            ui_type = "UI_BUTTON"

        # 3. Xử lý đa nhãn (Multi-key Labeling) cho nghiệp vụ vàng
        biz_labels = []
        for kw_name, kw_code in keywords_dict.items():
            if kw_name in clean_val:
                biz_labels.append(kw_code)

        # 4. Gom Metadata chuyên sâu để training AI Agent
        full_metadata = {
            'ui_context': {
                'type': ui_type,
                'is_merged': is_merged,
                'coordinate': cell.coordinate,
                'alignment': cell.alignment.horizontal if hasattr(cell.alignment, 'horizontal') else "left",
            },
            'business_context': {
                'labels': biz_labels,
                'is_formula': val_str.startswith('='),
                'data_type': type(cell.value).__name__, # Lưu lại tên kiểu gốc (ví dụ: 'time') để AI hiểu ngữ cảnh
                'raw_value': raw_val if not val_str.startswith('=') else None 
            },
            'style_context': {
                'background_color': bg_color,
                'is_bold': is_bold,
                'font_size': cell.font.size if hasattr(cell.font, 'size') else None,
                'number_format': cell.number_format if hasattr(cell, 'number_format') else "General"
            }
        }

        # 5. Khởi tạo đối tượng DataField (Sẵn sàng để bulk_create)
        # Lưu ý: DataField phải có trường 'raw_value' trong models.py
        from apps.excel_miner.models import DataField # Import cục bộ nếu cần
        return DataField(
            sheet=sheet_obj,
            cell_address=cell.coordinate,
            label=biz_labels[0] if biz_labels else None,
            value=val_str,
            raw_value=raw_val if not val_str.startswith('=') else None,
            formula=val_str if val_str.startswith('=') else "",
            comment=cell.comment.text if hasattr(cell.comment, 'text') else "",
            color_code=bg_color,
            field_type=formula_dtype if val_str.startswith('=') else default_dtype,
            metadata=full_metadata,
            confidence_score=1.0 if biz_labels else 0.5
        )
    
    class Meta:
        verbose_name = "Dự án Excel"
        verbose_name_plural = "Dự án Excel"

class ExcelSheet(models.Model):
    project = models.ForeignKey(ExcelProject, on_delete=models.CASCADE, related_name='sheets')
    name = models.CharField(max_length=255, verbose_name="Tên Sheet")
    sheet_index = models.IntegerField()
    # Thêm trường này để AI biết sheet này thuộc nhóm nghiệp vụ nào (Vàng 24k, Vàng Ý, Thợ...)
    category = models.CharField(max_length=100, null=True, blank=True) 
    
    class Meta:
        verbose_name = "Sheet Excel"
        verbose_name_plural = "Danh sách Sheets"

    def __str__(self): return self.name

class DataField(models.Model):
    sheet = models.ForeignKey(ExcelSheet, on_delete=models.CASCADE, related_name='fields')
    cell_address = models.CharField(max_length=10) 
    label = models.CharField(max_length=255, null=True, blank=True)
    
    # Value dùng để hiển thị/Chat (String)
    value = models.TextField(null=True, blank=True)
    
    # Raw Value dùng để tính toán (Lưu trữ đa kiểu dữ liệu: Float, Int, Date...)
    # Trong Django, dùng JSONField để lưu Raw Value là linh hoạt nhất
    raw_value = models.JSONField(null=True, blank=True, verbose_name="Giá trị gốc (Số/Ngày)")
    
    formula = models.TextField(null=True, blank=True)
    comment = models.TextField(null=True, blank=True, verbose_name="Ghi chú trong ô")
    metadata = models.JSONField(default=dict, blank=True, verbose_name="Phân tích chuyên sâu (AI)")

    field_type = models.ForeignKey(
        'system_monitor.DataType', 
        on_delete=models.SET_NULL, 
        null=True, 
        verbose_name="Loại dữ liệu"
    )
    
    confidence_score = models.FloatField(default=1.0)
    is_verified = models.BooleanField(default=False, verbose_name="Vũ đã xác nhận")
    color_code = models.CharField(max_length=20, null=True, blank=True)
    
    class Meta:
        verbose_name = "Trường dữ liệu"
        verbose_name_plural = "Dữ liệu chi tiết từ Sheets"
        unique_together = ('sheet', 'cell_address')

    def __str__(self):
        return f"{self.sheet.name} - {self.cell_address}"

    @property
    def get_real_value(self):
        """Ưu tiên lấy giá trị gốc để tính toán, nếu không có mới lấy value string"""
        return self.raw_value if self.raw_value is not None else self.value

    @property
    def ai_summary(self):
        biz_labels = self.metadata.get('business_context', {}).get('labels', [])
        label_desc = f" [{', '.join(biz_labels)}]" if biz_labels else ""
        
        # Hiển thị Raw Value nếu là số cho dễ nhìn trong Admin
        display_val = self.raw_value if self.raw_value is not None else self.value

        if self.comment:
            return f"📌 {self.comment[:30]}...{label_desc}"
        if self.formula:
            return f"🧮 {self.formula[:30]}...{label_desc}"
        return f"📄 {str(display_val)[:30]}...{label_desc}"
    
    def get_business_description(self):
        ui = self.metadata.get('ui_context', {}).get('type', 'DATA_CELL')
        biz = self.metadata.get('business_context', {})
        labels = biz.get('labels', [])
        
        desc = f"Thành phần: {ui}"
        if labels:
            desc += f" | Nghiệp vụ: {', '.join(labels)}"
        if biz.get('is_formula'):
            desc += " | Có công thức tính toán"
        
        # Thêm thông tin về kiểu dữ liệu gốc cho AI
        if self.raw_value is not None:
            desc += f" | Kiểu dữ liệu: {type(self.raw_value).__name__}"
            
        return desc

    def save(self, *args, **kwargs):
        super().save(*args, **kwargs)