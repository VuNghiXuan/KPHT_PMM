import re
import openpyxl
import datetime
import traceback
from openpyxl.cell.cell import MergedCell

# Import các Model liên quan
from apps.system_monitor.models import DataType
from .models import ExcelSheet, DataField
from apps.ai_knowledge.models import KnowledgeDraft

class ExcelMinerService:
    def __init__(self):
        self.keywords_dict = {}
        self.important_codes = []
        self.default_dtype = None
        self.formula_dtype = None

    def _prepare_resources(self):
        """Tải cấu hình DataType từ DB để định nghĩa bộ lọc nghiệp vụ."""
        data_types = DataType.objects.all()
        self.keywords_dict = {
            dt.name.lower(): {
                'code': dt.code, 
                'is_important': getattr(dt, 'is_important', False)
            } for dt in data_types
        }
        self.important_codes = [v['code'] for k, v in self.keywords_dict.items() if v['is_important']]
        
        # Thiết lập mặc định nếu DB chưa có dữ liệu cấu hình
        if not self.important_codes:
            self.important_codes = ['AMOUNT', 'WEIGHT', 'PERSON_NAME', 'BANK_ACCOUNT']

        self.default_dtype, _ = DataType.objects.get_or_create(code='MOCK', defaults={'name': 'Dữ liệu thô'})
        self.formula_dtype, _ = DataType.objects.get_or_create(code='FORMULA', defaults={'name': 'Công thức'})

    def process_project(self, project):
        self._prepare_resources()
        try:
            # data_only=False để giữ lại công thức (logic tiệm vàng)
            wb = openpyxl.load_workbook(project.file_path.path, data_only=False, read_only=False)
            total_sheets = len(wb.sheetnames)
            
            seen_drafts = set() 
            all_drafts = []

            for index, sheet_name in enumerate(wb.sheetnames):
                ws = wb[sheet_name]
                
                # 1. Tạo/Cập nhật thông tin Sheet
                sheet_obj, _ = ExcelSheet.objects.update_or_create(
                    project=project, name=sheet_name,
                    defaults={'sheet_index': index, 'category': 'TAB_UI'}
                )

                # Xóa dữ liệu cũ để nạp mới (tránh rác)
                DataField.objects.filter(sheet=sheet_obj).delete()
                fields_to_create = []

                # 2. Lấy danh sách các vùng ô gộp để xử lý tiêu đề Form
                merged_cells_ranges = ws.merged_cells.ranges

                # Giới hạn vùng đọc để tránh đọc hàng triệu dòng trống
                max_r = min(ws.max_row, 2000) # Điều chỉnh tùy độ dài thực tế của anh
                max_c = min(ws.max_column, 60)

                for row in ws.iter_rows(max_row=max_r, max_col=max_c):
                    for cell in row:
                        # Kiểm tra ô gộp: Nếu là ô gộp, chỉ xử lý ô đầu tiên (top-left)
                        is_merged = False
                        for merged_range in merged_cells_ranges:
                            if cell.coordinate in merged_range:
                                if cell.coordinate != merged_range.start_cell.coordinate:
                                    is_merged = True # Bỏ qua các ô phụ trong vùng gộp
                                break
                        
                        if is_merged:
                            continue

                        has_comment = hasattr(cell, 'comment') and cell.comment
                        
                        # Chỉ xử lý nếu ô có giá trị hoặc có ghi chú/công thức
                        if cell.value is not None or has_comment:
                            field_obj = self._prepare_cell_logic(ws, sheet_obj, cell)
                            fields_to_create.append(field_obj)
                            
                            # --- LOGIC TẠO BẢN THẢO (Tri thức Agent) ---
                            # Chỉ tạo Draft cho Công thức hoặc Ghi chú (Nơi chứa logic/quy trình)
                            if field_obj.formula or has_comment:
                                # Key dựa trên nội dung để tránh trùng lặp cùng một logic
                                draft_key = f"{sheet_name}_{field_obj.formula}_{str(has_comment)}"
                                if draft_key not in seen_drafts:
                                    draft = self._create_knowledge_draft(project, field_obj, cell)
                                    if draft: 
                                        all_drafts.append(draft)
                                        seen_drafts.add(draft_key)
                    
                    # Bulk create theo từng dòng lớn để tránh tràn RAM
                    if len(fields_to_create) >= 1000:
                        DataField.objects.bulk_create(fields_to_create)
                        fields_to_create = []

                # Lưu số còn lại của sheet
                if fields_to_create:
                    DataField.objects.bulk_create(fields_to_create)
            
            # 3. Lưu toàn bộ bản thảo tri thức để anh duyệt
            if all_drafts:
                KnowledgeDraft.objects.bulk_create(all_drafts, ignore_conflicts=True)
            
            return True, f"Thành công: {total_sheets} sheets, {len(all_drafts)} logic mới được tìm thấy."
            
        except Exception:
            return False, traceback.format_exc()

    def _create_knowledge_draft(self, project, field_obj, cell):
        """Đã tối ưu: Ép rớt dòng, chống lỗi NoneType và khớp Model thực tế"""
        # 1. Kiểm tra an toàn dữ liệu đầu vào
        if not project or not field_obj:
            return None

        comment = field_obj.metadata['ui_context'].get('comment', '')
        
        # 2. Phân loại tri thức cho tiệm vàng
        if field_obj.metadata['schema_mapping'].get('is_cross_sheet'):
            cat_label = '[HỆ THỐNG]'
        elif field_obj.formula:
            cat_label = '[LOGIC]'
        else:
            cat_label = '[THUẬT NGỮ]'

        title = field_obj.metadata['area_context'].get('smart_label') or f"Logic_{field_obj.cell_address}"
        
        # 3. Gom nội dung vào trường 'content' (Dùng cho giao diện ép rớt dòng)
        content_body = f"{cat_label} {title}\n"
        content_body += f"Vị trí: {field_obj.sheet.name}!{field_obj.cell_address}\n"
        
        if field_obj.formula:
            content_body += f"Công thức: {field_obj.formula}\n"
            deps = field_obj.metadata['schema_mapping'].get('depends_on_sheets', [])
            if deps:
                content_body += f"Liên kết từ: {', '.join(deps)}\n"
        
        if comment:
            content_body += f"Ghi chú tiệm vàng: {comment}\n"

        # 4. TRẢ VỀ OBJECT (Đảm bảo đúng tên trường trong DB)
        try:
            return KnowledgeDraft(
                project_id=project.id,    # Sử dụng .id để tránh AttributeError pk
                content=content_body,     # Nội dung đã gom nhóm
                status='PENDING',
                version=1
            )
        except Exception as e:
            print(f"Lỗi khởi tạo Draft: {str(e)}")
            return None

    def _prepare_cell_logic(self, ws, sheet_obj, cell):
        """Phân tích đa tầng nâng cao: Thêm Schema Mapping và Data Flow."""
        is_merged = isinstance(cell, MergedCell)
        raw_val = cell.value if not is_merged else None 
        
        if isinstance(raw_val, (datetime.datetime, datetime.date, datetime.time)):
            raw_val = raw_val.isoformat()

        val_str = str(raw_val).strip() if raw_val is not None else ""
        clean_val = " ".join(val_str.lower().split())
        is_formula = val_str.startswith('=')

        # --- 1. THU THẬP NGỮ CẢNH XUNG QUANH ---
        neighbor_left = ws.cell(row=cell.row, column=max(1, cell.column - 1)).value if cell.column > 1 else None
        neighbor_top = ws.cell(row=max(1, cell.row - 1), column=cell.column).value if cell.row > 1 else None
        col_header = ws.cell(row=1, column=cell.column).value or ws.cell(row=2, column=cell.column).value
        
        cell_comment = cell.comment.text if hasattr(cell, 'comment') and cell.comment else ""
        smart_label = self._get_smart_label(ws, cell)

        # --- 2. PHÂN LOẠI BUSINESS LABEL & ENTITY UNIFICATION ---
        biz_labels = []
        sources = [clean_val, str(col_header).lower(), str(neighbor_left).lower(), str(neighbor_top).lower(), smart_label.lower()]
        for src in sources:
            if not src or src == 'none': continue
            for kw, info in self.keywords_dict.items():
                if kw in src:
                    code = info['code'] if isinstance(info, dict) else info
                    if code not in biz_labels: biz_labels.append(code)
            if biz_labels: break

        # --- 3. PHÂN TÍCH LUỒNG DỮ LIỆU (DATA FLOW) TỪ CÔNG THỨC ---
        referenced_sheets = []
        if is_formula:
            # Regex bắt tên Sheet trong công thức (ví dụ: 'Sheet 1'!A1 hoặc Sheet1!A1)
            referenced_sheets = re.findall(r"['\"]?([^'\"!]+)['\"]?!", val_str)
            referenced_sheets = list(set(referenced_sheets)) # Loại bỏ trùng lặp

        # --- 4. NHẬN DIỆN STYLE & MÀU SẮC (UI RECOGNITION) ---
        bg_color = "00000000"
        if hasattr(cell.fill, 'start_color'):
            bg_color = str(cell.fill.start_color.index) if cell.fill.start_color.type == 'indexed' else str(cell.fill.start_color.rgb)
        
        is_bold = cell.font.bold if hasattr(cell.font, 'bold') else False
        
        ui_type = "DATA_CELL"
        if is_bold and bg_color not in ["00000000", "FFFFFFFF", "N/A"]: 
            ui_type = "GRID_HEADER"
        elif bg_color not in ["00000000", "FFFFFFFF", "N/A"] and 0 < len(val_str) < 30: 
            ui_type = "UI_BUTTON"

        # --- 5. PHÂN NHÓM CHỨC NĂNG ---
        func_group = "GENERAL_INFO"
        if ui_type == "UI_BUTTON": 
            func_group = "ACTION_TRIGGER"
        elif ui_type == "GRID_HEADER": 
            func_group = "SECTION_HEADER"
        elif any(label in self.important_codes for label in biz_labels) or is_formula:
            func_group = "CRITICAL_INPUT_VALIDATION"
        elif biz_labels: 
            func_group = "DATA_INPUT_FIELD"

        return DataField(
            sheet=sheet_obj, 
            cell_address=cell.coordinate,
            label=biz_labels[0] if biz_labels else None,
            value=val_str, 
            raw_value=raw_val if not is_formula else None,
            formula=val_str if is_formula else "",
            color_code=bg_color,
            field_type=self.formula_dtype if is_formula else self.default_dtype,
            metadata={
                'ui_context': {
                    'type': ui_type, 
                    'functional_group': func_group, 
                    'coordinate': cell.coordinate,
                    'comment': cell_comment
                },
                'area_context': {
                    'neighbor_left': str(neighbor_left), 
                    'neighbor_top': str(neighbor_top), 
                    'parent_col_title': str(col_header),
                    'smart_label': smart_label
                },
                'business_context': {
                    'labels': biz_labels, 
                    'is_formula': is_formula,
                    'is_required': bg_color in ["FFFF0000", "FFFFC000"],
                    'entity_type': biz_labels[0] if biz_labels else "UNKNOWN" # Thống nhất thực thể
                },
                'schema_mapping': {
                    'data_flow_direction': 'INBOUND' if is_formula else 'SOURCE',
                    'depends_on_sheets': referenced_sheets, # Các sheet cung cấp dữ liệu cho ô này
                    'is_cross_sheet': len(referenced_sheets) > 0
                }
            }
        )

    def _create_knowledge_draft(self, project, field_obj, cell):
        comment = field_obj.metadata['ui_context'].get('comment', '')
        
        # Phân loại tri thức dựa trên logic đã bóc tách
        if field_obj.metadata['schema_mapping']['is_cross_sheet']:
            category = 'SYSTEM_INTEGRATION'
        elif field_obj.formula:
            category = 'LOGIC'
        else:
            category = 'TERM'

        # Nhãn hiển thị chính
        display_term = field_obj.metadata['area_context'].get('smart_label') or f"Logic_{field_obj.cell_address}"
        
        # Nội dung chi tiết cho AI
        definition = f"Vị trí: {field_obj.sheet.name}!{field_obj.cell_address}\n"
        if field_obj.formula:
            definition += f"Công thức: {field_obj.formula}\n"
        if comment:
            definition += f"Ghi chú: {comment}\n"

        # TRUYỀN ĐÚNG CÁC TRƯỜNG TRONG MODEL (Sửa lỗi TypeError)
        return KnowledgeDraft(
            project_id=project.id,    # Dùng project_id thay vì project
            term=display_term[:255],  # Dùng term thay vì title
            category=category,
            content=definition,       # Dùng content thay vì definition
            origin_metadata=field_obj.metadata,
            status='PENDING'
        )

    def _get_smart_label(self, ws, cell):
        """Tìm nhãn văn bản gần nhất để định danh ô trên giao diện Form."""
        # 1. Thử lấy nhãn bên trái (thường dùng cho Form Input)
        left_val = ws.cell(row=cell.row, column=max(1, cell.column - 1)).value
        if left_val and isinstance(left_val, str) and len(left_val) < 50:
            return str(left_val).strip()
        
        # 2. Thử lấy nhãn phía trên (thường dùng cho Header Table)
        top_val = ws.cell(row=max(1, cell.row - 1), column=cell.column).value
        if top_val and isinstance(top_val, str) and len(top_val) < 50:
            return str(top_val).strip()
            
        return ""