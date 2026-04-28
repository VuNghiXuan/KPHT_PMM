import openpyxl
import os
import json
from database.models import ExcelSheet, DataGroup, DataField
from config import Config

class ExcelMiner:
    def __init__(self, file_path):
        self.file_path = os.path.abspath(file_path)
        try:
            # Lấy công thức thô và giá trị đối soát
            self.wb = openpyxl.load_workbook(self.file_path, data_only=False)
            self.wb_val = openpyxl.load_workbook(self.file_path, data_only=True)
        except Exception as e:
            raise Exception(f"Lỗi đọc file: {e}")

    def _get_color(self, cell) -> str:
        if cell.fill and hasattr(cell.fill, 'start_color'):
            return str(cell.fill.start_color.index)
        return "N/A"

    def _get_merged_value(self, ws_val, cell):
        """Xử lý lấy giá trị cho ô nằm trong vùng bị gộp (Merged Cells)"""
        for merged_range in ws_val.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # Trả về giá trị của ô đầu tiên trong vùng gộp
                return ws_val.cell(merged_range.min_row, merged_range.min_col).value
        return None

    def scan_project(self) -> list:
        project_data = []
        # Cấu trúc lại theo phân cấp Sheet để AI dễ đọc mục lục
        knowledge_backup = {} 

        for name in self.wb.sheetnames:
            ws = self.wb[name]
            ws_val = self.wb_val[name]
            
            sheet_obj = ExcelSheet(sheet_name=name, status="SCANNING")
            raw_group = DataGroup(group_name=f"Full_Data_{name}")
            knowledge_backup[name] = []
            
            print(f"🔍 Đang 'vét cạn' tri thức tại Sheet: [{name}]")

            # Lấy vùng dữ liệu thực tế để tránh quét hàng triệu ô trống của Excel
            dim = ws.calculate_dimension() 
            
            for row in ws[dim]:
                for cell in row:
                    # 1. Xử lý giá trị (Ưu tiên ô gộp nếu ô hiện tại rỗng)
                    real_val = ws_val[cell.coordinate].value
                    if real_val is None:
                        real_val = self._get_merged_value(ws_val, cell)
                    
                    raw_val = cell.value
                    has_comment = cell.comment is not None
                    color = self._get_color(cell)
                    has_color = color not in ["00000000", "N/A", "0"]
                    
                    # Điều kiện vét: Có dữ liệu, có công thức, có màu hoặc có ghi chú
                    if raw_val is not None or has_comment or has_color:
                        val_str = str(raw_val).strip()
                        is_formula = val_str.startswith('=')
                        
                        # --- AI LOGIC TAGGING ---
                        f_type = "DATA"
                        is_bold = cell.font.bold if cell.font else False
                        
                        # Nhận diện Button (Dựa trên từ khóa và độ dài chuỗi)
                        action_keywords = getattr(Config, 'ACTION_KEYWORDS', ['LƯU', 'IN', 'XÓA', 'TÍNH'])
                        if any(k in val_str.upper() for k in action_keywords) and len(val_str) < 15:
                            f_type = "UI_BUTTON"
                        
                        # Nhận diện Header (Chữ đậm hoặc có màu nền)
                        elif is_bold and not is_formula:
                            f_type = "GRID_HEADER"
                        
                        # Nhận diện ô tính toán quan trọng (Có màu nền)
                        elif has_color and is_formula:
                            f_type = "CALC_CELL"

                        field_info = {
                            "coord": cell.coordinate,
                            "value": str(real_val) if real_val is not None else "",
                            "formula": val_str if is_formula else None,
                            "comment": cell.comment.text.strip() if has_comment else None,
                            "color": color,
                            "is_bold": is_bold,
                            "type": f_type,
                            "is_hidden": ws.row_dimensions[cell.row].hidden or ws.column_dimensions[cell.column_letter].hidden
                        }

                        # Lưu vào DB Object
                        field_obj = DataField(
                            coord=field_info["coord"],
                            row=cell.row,
                            column=cell.column,
                            label=f"{f_type} (NOTE: {field_info['comment']})" if has_comment else f_type,
                            value=field_info["value"],
                            formula=field_info["formula"],
                            color_code=field_info["color"],
                            field_type=f_type
                        )

                        raw_group.fields.append(field_obj)
                        knowledge_backup[name].append(field_info)

            if raw_group.fields:
                sheet_obj.groups.append(raw_group)
                project_data.append(sheet_obj)

        self._save_backup(knowledge_backup)
        return project_data

    def _save_backup(self, data):
        backup_path = "knowledge_backup.json"
        with open(backup_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"💾 Gen hệ thống đã được đúc vào: {backup_path}")

    def close(self):
        self.wb.close()
        self.wb_val.close()